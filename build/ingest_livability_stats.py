#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""自治体の「暮らし実態」統計を公開オープンデータ（キー不要のCSV/Excel）から取り込み、
別統計DB data/livability_stats.db に蓄積する（reinfolib_tokyo.db と同じ「別DB→build時結合」方式）。

制度データ（gov_life_support.sqlite3）とは独立。municipality_code / municipality_name で突合。
long形式（1指標1行）なので指標の追加が容易。

指標(indicator)例:
  taikijido           … 待機児童数（こども家庭庁 保育所等関連状況取りまとめ）
  hoiku_riyou_rate    … 保育サービス利用率
  （今後）population / zaiseiryoku_index / ...

使い方:
  python3 build/ingest_livability_stats.py            # 既定ソースを取込
出典ファイルは curl で取得（HTTPS_PROXY 経由）。取得日・出典URLを行に記録する。
"""
import os
import sqlite3
import subprocess
import sys
import tempfile

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GOVDB = os.path.join(ROOT, "gov_life_support.sqlite3")
STATSDB = os.path.join(ROOT, "data", "livability_stats.db")
FETCHED = os.environ.get("FETCHED_DATE", "2026-07-30")

SCHEMA = """
CREATE TABLE IF NOT EXISTS municipality_stats(
  municipality_code TEXT,
  municipality_name TEXT,
  indicator   TEXT,      -- 指標キー
  value       REAL,      -- 数値
  unit        TEXT,      -- 単位（人 / % など）
  year        TEXT,      -- 基準時点（例 2025-04-01）
  source_name TEXT,
  source_url  TEXT,
  fetched_at  TEXT,
  PRIMARY KEY(municipality_name, indicator, year)
);
"""


def curl(url, dest):
    subprocess.run(["curl", "-sS", "-L", "--max-time", "90", "-o", dest, url], check=True)
    if not os.path.getsize(dest):
        raise RuntimeError(f"empty download: {url}")


def name2code(c):
    return {n: code for n, code in c.execute(
        "SELECT municipality_name, municipality_code FROM municipalities").fetchall()}


def upsert(c, rows):
    c.executemany(
        "INSERT OR REPLACE INTO municipality_stats"
        "(municipality_code,municipality_name,indicator,value,unit,year,source_name,source_url,fetched_at)"
        " VALUES(?,?,?,?,?,?,?,?,?)", rows)


def ingest_taikijido(gov, stats):
    """こども家庭庁『保育所等関連状況取りまとめ（令和7年4月1日）表4 区市町村別』(東京都分)。"""
    import openpyxl
    import re
    url = "https://www.metro.tokyo.lg.jp/documents/d/tosei/20250829_17_04"
    src = "こども家庭庁/東京都 保育サービスの状況 表4（令和7年4月1日）"
    year = "2025-04-01"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    curl(url, path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    n2c = name2code(gov)
    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        name = row[1]
        if not isinstance(name, str):
            continue
        name = name.strip()
        if not re.search(r"(区|市|町|村)$", name) or name not in n2c:
            continue
        pre, wait, rate = row[2], row[5], row[4]
        code = n2c[name]
        if pre is not None:
            rows.append((code, name, "shugakumae_jinko", float(pre), "人", year, src, url, FETCHED))
        if wait is not None:
            rows.append((code, name, "taikijido", float(wait), "人", year, src, url, FETCHED))
        if rate is not None:
            rows.append((code, name, "hoiku_riyou_rate", round(float(rate) * 100, 1), "%", year, src, url, FETCHED))
    os.unlink(path)
    upsert(stats, rows)
    return len({r[1] for r in rows}), len(rows)


def ingest_population(gov, stats):
    """東京都『住民基本台帳による世帯と人口（令和7年1月）第5表』CSV（町丁別）。
    町丁別地域階層=0 の行が各自治体の合計。人口総数・世帯数を取り込む。"""
    import csv as _csv
    url = "https://www.toukei.metro.tokyo.lg.jp/juukiy/2025/jy25qv0500.csv"
    page = "https://www.toukei.metro.tokyo.lg.jp/juukiy/2025/jy25q10501.htm"
    src = "東京都 住民基本台帳による世帯と人口 令和7年1月 第5表"
    year = "2025-01-01"
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tf:
        path = tf.name
    curl(url, path)
    n2c = name2code(gov)
    rows = []
    with open(path, encoding="utf-8") as f:
        for r in _csv.reader(f):
            if len(r) < 6 or r[2].strip() != "0":
                continue
            name = r[1].replace("総数", "").strip()
            if name not in n2c:
                continue
            try:
                pop, setai = int(r[5]), int(r[4])
            except ValueError:
                continue
            code = n2c[name]
            rows.append((code, name, "population", float(pop), "人", year, src, page, FETCHED))
            rows.append((code, name, "setai", float(setai), "世帯", year, src, page, FETCHED))
    os.unlink(path)
    upsert(stats, rows)
    return len({r[1] for r in rows}), len(rows)


def ingest_age_structure(gov, stats):
    """東京都『住民基本台帳による東京都の世帯と人口 令和7年1月』第3-1表
    区市町村・年齢3区分別人口（人口総数）CSV。地域階層=4 の行が各自治体。
    年少(0-14)/生産年齢(15-64)/老年(65-) の総数から高齢化率・年少人口率を算出。"""
    import csv as _csv
    url = "https://www.toukei.metro.tokyo.lg.jp/juukiy/2025/jy25qv0301.csv"
    page = "https://www.toukei.metro.tokyo.lg.jp/juukiy/2025/jy25000001.htm"
    src = "東京都 住民基本台帳による東京都の世帯と人口 令和7年1月 第3-1表"
    year = "2025-01-01"
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tf:
        path = tf.name
    curl(url, path)
    n2c = name2code(gov)
    rows = []
    with open(path, encoding="utf-8-sig") as f:  # BOM付きUTF-8
        for r in _csv.reader(f):
            if len(r) < 10 or r[0].strip() != "4":  # 地域階層=4 が区市町村
                continue
            name = r[2].strip()
            if name not in n2c:
                continue
            def num(x):
                try:
                    return int(str(x).replace(",", "").strip())
                except ValueError:
                    return None
            young, work, old = num(r[3]), num(r[6]), num(r[9])
            if None in (young, work, old):
                continue
            total = young + work + old
            if total <= 0:
                continue
            code = n2c[name]
            rows.append((code, name, "koreika_rate", round(100 * old / total, 1), "%", year, src, page, FETCHED))
            rows.append((code, name, "nenshou_rate", round(100 * young / total, 1), "%", year, src, page, FETCHED))
    os.unlink(path)
    upsert(stats, rows)
    return len({r[1] for r in rows}), len(rows)


def main():
    # 決定的に再生成（既存statsDBを作り直す）
    if os.path.exists(STATSDB):
        os.remove(STATSDB)
    gov = sqlite3.connect(GOVDB).cursor()
    sdb = sqlite3.connect(STATSDB)
    st = sdb.cursor()
    st.executescript(SCHEMA)
    m1, c1 = ingest_taikijido(gov, st)
    m2, c2 = ingest_population(gov, st)
    m3, c3 = ingest_age_structure(gov, st)
    munis, cells = max(m1, m2, m3), c1 + c2 + c3
    print(f"人口/世帯: {m2}自治体 / {c2}セル")
    print(f"年齢構成(高齢化率/年少人口率): {m3}自治体 / {c3}セル")
    sdb.commit()
    tot = st.execute("SELECT COUNT(*) FROM municipality_stats").fetchone()[0]
    inds = st.execute("SELECT indicator,COUNT(*) FROM municipality_stats GROUP BY indicator").fetchall()
    print(f"待機児童/保育: {munis}自治体 / {cells}セル 取込")
    print(f"municipality_stats 総行数={tot}  指標内訳={dict(inds)}")
    sdb.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
